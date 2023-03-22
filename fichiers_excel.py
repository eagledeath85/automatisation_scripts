# FICHIERS EXCEL
# .XLSX
# openpyxl

import openpyxl

# Ouvrir le fichier excel et recuperer son contenu
# La cle data_only permet de recuperer uniquement les valeurs contenues dans chaque cellule, et non les formules
workbook_oct = openpyxl.load_workbook('octobre.xlsx', data_only=True)
workbook_nov = openpyxl.load_workbook('novembre.xlsx', data_only=True)
workbook_dec = openpyxl.load_workbook('decembre.xlsx', data_only=True)

# Recuperer le nom de toutes les feuilles du workbook. Le resultat est une liste de str contenant le nom de chaque feuille
print(workbook_oct.sheetnames)

# Recuperer le contenu de Feuil1
sheet1 = workbook_oct[workbook_oct.sheetnames[0]]

# Recuperer le contenu de la cellule A1 via la methode cell(row, column)
# Une cellule vide retourne la valeur None
for row1 in range(2, 7):
    cell = sheet1.cell(row1, 2)
    print(cell.value)

print()


# Construction de la structure de donnees selon le format
# {"Pommes": [760, 600, 900]}


def ajouter_data_depuis_workbook(wb, d: dict):
    # Recuperer la feuille active
    sheet = wb.active

    # Afficher le contenu de la premiere colonne
    for row in range(2, sheet.max_row):
        nom_article = sheet.cell(row, 1).value
        if not nom_article:
            break
        total_ventes = sheet.cell(row, 4).value  # Recuperer la valeur correspondante dans la colonne 4
        # Verifier si la cle existe deja. Si oui, on ajoute la nvelle valeur total_vente a la liste [total_ventes]
        if d.get(nom_article):
            d[nom_article].append(total_ventes)
        else:
            d[nom_article] = [total_ventes]


donnees = {}
ajouter_data_depuis_workbook(workbook_oct, donnees)
ajouter_data_depuis_workbook(workbook_nov, donnees)
ajouter_data_depuis_workbook(workbook_dec, donnees)
print(donnees)

# Creer un workbook de sortie
workbook_sortie = openpyxl.Workbook()
# Creer un feuille active dans le workbook
sheet = workbook_sortie.active
# Remplir les cellules de titre a la main
sheet["A1"] = "Article"
sheet["B1"] = "Octobre"
sheet["C1"] = "Novembre"
sheet["D1"] = "Decembre"


# On demarre la liste a la ligne 2
row = 2
for i in donnees.items():
    nom_article = i[0]
    ventes = i[1]
    sheet.cell(row, 1).value = nom_article
    for j in range(0, len(ventes)):
        sheet.cell(row, 2+j).value = ventes[j]
    row += 1


# Sauvegarder le workbbok de sortie dans un fichier de sortie
workbook_sortie.save("recap_ventes_trimestre_4.xlsx")